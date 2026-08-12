"""XLSX-backed scheme storage for grid-planning parameter maintenance."""

from __future__ import annotations

import re
import unicodedata
import zipfile
import gc
import json
import shutil
import sqlite3
import tempfile
from copy import deepcopy
from dataclasses import dataclass
from datetime import datetime
from io import BytesIO
from pathlib import Path, PurePosixPath
from typing import Any
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook

import file_cache
import file_ops


WEB_ROOT = Path(__file__).resolve().parent
DEFAULT_SCHEME_ROOT = WEB_ROOT / "planning_schemes"
WORKBOOK_NAME = "parameters.xlsx"
TIME_SERIES_WORKBOOK_NAME = "time_series.xlsx"
SCHEME_META_NAME = "scheme_meta.json"
TIME_SERIES_SQLITE_NAME = ".time_series.sqlite3"
TIME_SERIES_SQLITE_SCHEMA_VERSION = "1"
SCHEME_ARCHIVE_MAX_FILES = 2000
SCHEME_ARCHIVE_MAX_UNCOMPRESSED_BYTES = 512 * 1024 * 1024
WORKBOOK_XML = "xl/workbook.xml"
WORKBOOK_RELS_XML = "xl/_rels/workbook.xml.rels"
MAIN_XML_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_XML_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PACKAGE_REL_XML_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
PARAMETER_WORKBOOK_CACHE = file_cache.FileCache("parameter_workbook", max_entries=96)
SHEET_ROW_COUNT_CACHE = file_cache.FileCache("sheet_row_count", max_entries=96)

# Sheet order and column order are part of the on-disk contract. The browser,
# workbook reader, and workbook writer all depend on this exact schema.
SHEET_SPECS: dict[str, tuple[str, list[str]]] = {
    "time_series": ("8760时序数据", ["hour_index", "datetime", "wind_speed", "solar_irradiance", "load", "temperature"]),
    "diesel_generators": (
        "柴发参数",
        [
            "name",
            "cost",
            "capacity",
            "power_upper",
            "power_lower",
            "fuel_rate",
            "inertia_constant_h",
            "primary_frequency_coefficient_k",
            "damping_coefficient_d",
            "governor_time_constant_t",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "wind_turbines": (
        "风机参数",
        [
            "name",
            "cost",
            "capacity",
            "cut_in_wind_speed",
            "rated_wind_speed",
            "cut_out_wind_speed",
            "is_grid_forming",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "photovoltaics": (
        "光伏参数",
        [
            "name",
            "cost",
            "capacity",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "storage_pcs": (
        "储能PCS参数",
        [
            "name",
            "cost",
            "power_capacity",
            "storage_charge_efficiency",
            "storage_discharge_efficiency",
            "is_grid_forming",
            "storage_equivalent_inertia_constant_h",
            "storage_equivalent_primary_frequency_coefficient_k",
            "storage_equivalent_damping_coefficient_d",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "storage_battery_packs": (
        "储能电池组参数",
        [
            "name",
            "cost",
            "battery_capacity",
            "soc_upper",
            "soc_lower",
            "self_discharge_rate",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "hydrogen_electrolyzers": (
        "电制氢参数",
        [
            "name",
            "cost",
            "power_capacity",
            "power_lower",
            "electric_to_hydrogen_efficiency",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "hydrogen_tanks": (
        "储氢罐参数",
        [
            "name",
            "cost",
            "hydrogen_tank_capacity",
            "soc_upper",
            "soc_lower",
            "self_discharge_rate",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "fuel_cells": (
        "燃料电池参数",
        [
            "name",
            "cost",
            "power_capacity",
            "hydrogen_to_electric_efficiency",
            "quantity_lower",
            "quantity_upper",
            "design_life_years",
        ],
    ),
    "planning_parameters": (
        "规划参数",
        [
            "diesel_price",
            "diesel_minimum_on_hours",
            "diesel_minimum_off_hours",
            "operation_mode",
            "winter_start_month",
            "winter_start_day",
            "winter_end_month",
            "winter_end_day",
            "green_power_ratio_lower",
            "optimization_time_limit_minutes",
            "preferred_solver",
            "initial_storage_soc_ratio",
            "storage_balance_mode",
            "initial_hydrogen_storage_ratio",
            "post_disturbance_power_balance_enabled",
            "renewable_n_1_enabled",
            "renewable_disturbance_enabled",
            "load_disturbance_enabled",
            "load_up_disturbance_factor",
            "load_down_disturbance_factor",
            "renewable_down_disturbance_factor",
            "frequency_security_constraint_enabled",
            "nominal_frequency_hz",
            "frequency_nadir_lower_hz",
            "frequency_peak_upper_hz",
            "frequency_lower_security_margin_hz",
            "frequency_upper_security_margin_hz",
            "load_frequency_coefficient_d",
            "rocof_upper_hz_per_s",
            "steady_state_frequency_lower_hz",
            "steady_state_frequency_upper_hz",
            "frequency_governor_time_constant_s",
            "frequency_nadir_evaluation_duration_s",
            "nadir_linearization_samples_per_axis",
            "nadir_linearization_interval_ratio",
            "network_synchronization_coefficient_base",
            "network_synchronization_coefficient_slope",
            "network_synchronization_reference_load_kw",
            "storage_frequency_regulation_enabled",
        ],
    ),
}

# Default rows seed a new scheme and also supply fallback values for older
# workbooks that are missing newly added columns.
DEFAULT_DEVICE_ROWS: dict[str, list[dict[str, Any]]] = {
    "diesel_generators": [
        {
            "name": "柴发1",
            "capacity": 100,
            "cost": 0,
            "power_upper": 100,
            "power_lower": 20,
            "fuel_rate": 0.26,
            "inertia_constant_h": 3.5,
            "primary_frequency_coefficient_k": 0.4,
            "damping_coefficient_d": 0.01,
            "governor_time_constant_t": 0.6,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "wind_turbines": [
        {
            "name": "风机1",
            "capacity": 50,
            "cost": 0,
            "cut_in_wind_speed": 3,
            "rated_wind_speed": 12,
            "cut_out_wind_speed": 25,
            "is_grid_forming": 0,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "photovoltaics": [
        {
            "name": "光伏1",
            "capacity": 50,
            "cost": 0,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "storage_pcs": [
        {
            "name": "储能PCS1",
            "power_capacity": 50,
            "storage_charge_efficiency": 0.95,
            "storage_discharge_efficiency": 0.95,
            "cost": 0,
            "quantity_lower": 0,
            "quantity_upper": 0,
            "is_grid_forming": 0,
            "storage_equivalent_inertia_constant_h": 2.5,
            "storage_equivalent_primary_frequency_coefficient_k": 0.5,
            "storage_equivalent_damping_coefficient_d": 0.05,
        }
    ],
    "storage_battery_packs": [
        {
            "name": "储能电池组1",
            "battery_capacity": 200,
            "soc_upper": 0.9,
            "soc_lower": 0.1,
            "self_discharge_rate": 0.01,
            "cost": 0,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "hydrogen_electrolyzers": [
        {
            "name": "电制氢1",
            "power_capacity": 50,
            "power_lower": 0,
            "cost": 0,
            "electric_to_hydrogen_efficiency": 0.7,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "hydrogen_tanks": [
        {
            "name": "储氢罐1",
            "hydrogen_tank_capacity": 100,
            "soc_upper": 0.85,
            "soc_lower": 0.15,
            "self_discharge_rate": 0.001,
            "cost": 0,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
    "fuel_cells": [
        {
            "name": "燃料电池1",
            "power_capacity": 50,
            "cost": 0,
            "hydrogen_to_electric_efficiency": 0.55,
            "quantity_lower": 0,
            "quantity_upper": 0,
        }
    ],
}

# Planning parameters live in a single global row instead of repeating per
# device family.
DEFAULT_PLANNING_PARAMETERS: dict[str, Any] = {
    "diesel_price": 0,
    "diesel_minimum_on_hours": 12,
    "diesel_minimum_off_hours": 12,
    "operation_mode": "annual",
    "winter_start_month": 10,
    "winter_start_day": 1,
    "winter_end_month": 4,
    "winter_end_day": 30,
    "green_power_ratio_lower": 0,
    "optimization_time_limit_minutes": 60,
    "preferred_solver": "auto",
    "initial_storage_soc_ratio": 0.5,
    "storage_balance_mode": "daily",
    "initial_hydrogen_storage_ratio": 0.5,
    "storage_frequency_regulation_enabled": 0,
    "renewable_disturbance_enabled": 0,
    "load_up_disturbance_factor": 0,
    "load_down_disturbance_factor": 0,
    "renewable_down_disturbance_factor": 0,
    "frequency_security_constraint_enabled": 0,
    "nominal_frequency_hz": 50.0,
    "frequency_nadir_lower_hz": 49.5,
    "frequency_peak_upper_hz": 50.5,
    "frequency_lower_security_margin_hz": 0.0,
    "frequency_upper_security_margin_hz": 0.0,
    "load_frequency_coefficient_d": 0.0,
    "rocof_upper_hz_per_s": 1.0,
    "steady_state_frequency_lower_hz": 49.5,
    "steady_state_frequency_upper_hz": 50.5,
    "frequency_governor_time_constant_s": 0.6,
    "frequency_nadir_evaluation_duration_s": 20.0,
    "nadir_linearization_samples_per_axis": 4,
    "nadir_linearization_interval_ratio": 0.5,
    "network_synchronization_coefficient_base": 1.0,
    "network_synchronization_coefficient_slope": 0.0,
    "network_synchronization_reference_load_kw": 0.0,
    "post_disturbance_power_balance_enabled": 1,
    "renewable_n_1_enabled": 0,
    "load_disturbance_enabled": 0,
}

PLANNING_BOOLEAN_FIELDS = {
    "storage_frequency_regulation_enabled",
    "frequency_security_constraint_enabled",
    "post_disturbance_power_balance_enabled",
    "renewable_n_1_enabled",
    "renewable_disturbance_enabled",
    "load_disturbance_enabled",
}

OPERATION_MODE_ALIASES = {
    "": "annual",
    "annual": "annual",
    "year": "annual",
    "全年": "annual",
    "全年运行": "annual",
    "度夏": "summer",
    "度夏运行": "summer",
    "summer": "summer",
    "summer_only": "summer",
    "summer-only": "summer",
}

FIELD_DEFAULTS: dict[str, Any] = {
    **DEFAULT_PLANNING_PARAMETERS,
    "design_life_years": 20,
    "inertia_constant_h": 3.5,
    "primary_frequency_coefficient_k": 0.4,
    "damping_coefficient_d": 0.01,
    "governor_time_constant_t": 0.6,
    "rated_wind_speed": 12,
    "is_grid_forming": 0,
    "storage_equivalent_inertia_constant_h": 2.5,
    "storage_equivalent_primary_frequency_coefficient_k": 0.5,
    "storage_equivalent_damping_coefficient_d": 0.05,
    "storage_charge_efficiency": 0.95,
    "storage_discharge_efficiency": 0.95,
    "soc_upper": 0.9,
    "soc_lower": 0.1,
    "self_discharge_rate": 0.01,
}

DEVICE_FIELD_RULES: dict[str, dict[str, Any]] = {
    "quantity_lower": {"integer": True, "non_negative": True, "message": "数量上下限必须为非负整数"},
    "quantity_upper": {"integer": True, "non_negative": True, "message": "数量上下限必须为非负整数"},
    "design_life_years": {"integer": True, "positive": True, "message": "设计年限(年）必须为正整数"},
    "cost": {"non_negative": True, "message": "成本(万元/台)必须为非负浮点数"},
    "capacity": {"positive": True, "message": "容量(kW)必须为正实数"},
    "power_capacity": {"positive": True, "message": "容量(kW)必须为正实数"},
    "storage_charge_efficiency": {"min": 0, "max": 1, "positive": True, "message": "充电效率(0.0-1.0)必须在0到1之间，且必须大于0"},
    "storage_discharge_efficiency": {"min": 0, "max": 1, "positive": True, "message": "放电效率(0.0-1.0)必须在0到1之间，且必须大于0"},
    "battery_capacity": {"positive": True, "message": "容量(kWh)必须为正实数"},
    "soc_upper": {"min": 0, "max": 1, "message": "SOC上限(0.0-1.0)必须在0到1之间"},
    "soc_lower": {"min": 0, "max": 1, "message": "SOC下限(0.0-1.0)必须在0到1之间"},
    "self_discharge_rate": {"min": 0, "max": 0.01, "message": "自损耗率(0-1%/天)必须在0到0.01之间"},
    "is_grid_forming": {"integer": True, "min": 0, "max": 1, "message": "是否构网必须为0或1"},
    "storage_equivalent_inertia_constant_h": {"min": 0.0, "max": 20.0, "message": "等效惯量常数H(s)必须在0到20.0之间"},
    "storage_equivalent_primary_frequency_coefficient_k": {"min": 0.0, "max": 10.0, "message": "等效一次调频系数K必须在0到10.0之间"},
    "storage_equivalent_damping_coefficient_d": {"min": 0.0, "max": 20.0, "message": "等效阻尼系数D必须在0到20.0之间"},
    "hydrogen_tank_capacity": {"positive": True, "message": "容量(Nm3)必须为正实数"},
    "electric_to_hydrogen_efficiency": {"positive": True, "message": "电-氢效率(Nm3/kWh)必须为正实数"},
    "hydrogen_to_electric_efficiency": {"positive": True, "message": "氢-电效率(kWh/Nm3)必须为正实数"},
    "fuel_rate": {"positive": True, "message": "油耗率(kg/kWh)必须为正实数"},
    "inertia_constant_h": {"min": 0.0, "max": 20.0, "message": "惯量常数H(s)必须在0到20.0之间"},
    "primary_frequency_coefficient_k": {"min": 0.0, "max": 10.0, "message": "一次调频系数K必须在0到10.0之间"},
    "damping_coefficient_d": {"min": 0.0, "max": 20.0, "message": "阻尼系数D必须在0到20.0之间"},
    "governor_time_constant_t": {"min": 0.0001, "max": 20.0, "message": "调速时间常数T(s)必须在0.0001到20.0之间"},
    "power_lower": {"non_negative": True, "message": "功率下限(kW)必须为非负实数"},
    "cut_in_wind_speed": {"non_negative": True, "message": "切入风速(m/s)必须为非负实数"},
    "rated_wind_speed": {"positive": True, "message": "额定风速(m/s)必须为正实数"},
    "cut_out_wind_speed": {"non_negative": True, "message": "切出风速(m/s)必须为非负实数"},
}

INVALID_NAME_RE = re.compile(r'[<>:"/\\|?*]')


def sanitize_scheme_name(name: str) -> str:
    return "".join(
        char
        for char in str(name or "")
        if not char.isspace() and unicodedata.category(char) not in {"Cc", "Cf"}
    )


def validate_scheme_name(name: str) -> str:
    clean = sanitize_scheme_name(name)
    if clean in {"", ".", ".."} or INVALID_NAME_RE.search(clean) or ".." in clean:
        raise ValueError("方案名称不能为空，且不能包含路径或非法字符")
    return clean


def default_time_series() -> list[dict[str, Any]]:
    return [
        {
            "hour_index": hour,
            "datetime": f"H{hour:04d}",
            "wind_speed": 0,
            "solar_irradiance": 0,
            "load": 0,
            "temperature": 0,
        }
        for hour in range(1, 8761)
    ]


def default_payload(scheme: str) -> dict[str, Any]:
    # Build a complete in-memory scheme so the UI can render immediately after
    # a create action without special-casing missing sheets.
    payload: dict[str, Any] = {"scheme": scheme, "time_series": default_time_series(), "validation": []}
    for key in DEFAULT_DEVICE_ROWS:
        payload[key] = [with_field_defaults(row, SHEET_SPECS[key][1], key) for row in DEFAULT_DEVICE_ROWS[key]]
    payload["planning_parameters"] = [deepcopy(DEFAULT_PLANNING_PARAMETERS)]
    payload["capacity_limits"] = []
    return payload


def default_rows_for_key(key: str) -> list[dict[str, Any]]:
    if key == "time_series":
        return default_time_series()
    if key == "planning_parameters":
        return [deepcopy(DEFAULT_PLANNING_PARAMETERS)]
    return [with_field_defaults(row, SHEET_SPECS[key][1], key) for row in DEFAULT_DEVICE_ROWS.get(key, [])]


def field_default(header: str, fallback: Any = "") -> Any:
    return deepcopy(FIELD_DEFAULTS.get(header, fallback))


def field_default_for_key(key: str, header: str, fallback: Any = "") -> Any:
    if header == "self_discharge_rate":
        if key == "hydrogen_tanks":
            return 0.001
        if key == "storage_battery_packs":
            return 0.01
    if key == "hydrogen_tanks" and header == "soc_upper":
        return 0.85
    if key == "hydrogen_tanks" and header == "soc_lower":
        return 0.15
    return field_default(header, fallback)


def truthy_flag(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return float(value) != 0.0
    return str(value).strip().lower() in {"1", "true", "yes", "y", "是"}


def numeric_boolean_value(value: Any) -> int:
    return 1 if truthy_flag(value) else 0


def normalize_planning_parameter_row(row: dict[str, Any]) -> dict[str, Any]:
    normalized = with_field_defaults(row, SHEET_SPECS["planning_parameters"][1], "planning_parameters")
    for field in PLANNING_BOOLEAN_FIELDS:
        normalized[field] = numeric_boolean_value(normalized.get(field, 0))
    normalized["preferred_solver"] = normalize_preferred_solver(normalized.get("preferred_solver"))
    normalized["operation_mode"] = normalize_operation_mode(normalized.get("operation_mode"))
    for field in ("winter_start_month", "winter_start_day", "winter_end_month", "winter_end_day"):
        normalized[field] = int(max(1, round(numeric(normalized.get(field), DEFAULT_PLANNING_PARAMETERS[field]))))
    return normalized


def normalize_operation_mode(value: Any) -> str:
    mode = str(value or "annual").strip().lower()
    return OPERATION_MODE_ALIASES.get(mode, "annual")


def is_valid_operation_mode_value(value: Any) -> bool:
    return str(value or "").strip().lower() in OPERATION_MODE_ALIASES


def normalize_preferred_solver(value: Any) -> str:
    solver = str(value or "auto").strip().lower()
    aliases = {
        "": "auto",
        "automatic": "auto",
        "自动": "auto",
        "自动选择": "auto",
        "grb": "gurobi",
        "gurobi": "gurobi",
        "cplx": "cplex",
        "cplex": "cplex",
        "msk": "mosek",
        "mosek": "mosek",
        "highs": "scipy",
        "scipy": "scipy",
        "scipy highs": "scipy",
        "scipy-highs": "scipy",
    }
    return aliases.get(solver, "auto")


def with_field_defaults(row: dict[str, Any], headers: list[str], key: str = "") -> dict[str, Any]:
    # Fill gaps from the current schema so newer columns appear in older
    # workbooks without requiring a manual migration.
    normalized = deepcopy(row)
    for header in headers:
        if header not in normalized:
            normalized[header] = field_default_for_key(key, header, "")
    return normalized


def sanitize_payload_names(payload: dict[str, Any]) -> dict[str, Any]:
    for key in DEFAULT_DEVICE_ROWS:
        rows = payload.get(key)
        if not isinstance(rows, list):
            continue
        for row in rows:
            if isinstance(row, dict) and "name" in row:
                row["name"] = sanitize_scheme_name(row.get("name", ""))
    rows = payload.get("planning_parameters")
    if isinstance(rows, dict):
        payload["planning_parameters"] = [normalize_planning_parameter_row(rows)]
    elif isinstance(rows, list):
        payload["planning_parameters"] = [
            normalize_planning_parameter_row(row) if isinstance(row, dict) else normalize_planning_parameter_row({})
            for row in rows
        ]
    return payload


def normalize_shared_usernames(usernames: Any) -> list[str]:
    if isinstance(usernames, str):
        candidates = [usernames]
    elif isinstance(usernames, list):
        candidates = usernames
    else:
        candidates = []
    result: list[str] = []
    for item in candidates:
        username = str(item or "").strip()
        if username and username not in result:
            result.append(username)
    return result


def safe_archive_member_parts(filename: str) -> list[str]:
    raw = str(filename or "").replace("\\", "/")
    if "\x00" in raw:
        raise ValueError("方案压缩包包含非法路径")
    path = PurePosixPath(raw)
    if path.is_absolute():
        raise ValueError("方案压缩包包含非法路径")
    parts = [part for part in path.parts if part not in {"", "."}]
    if any(part == ".." for part in parts):
        raise ValueError("方案压缩包包含非法路径")
    return parts


def archive_strip_root_prefix(member_parts: list[list[str]]) -> str:
    if not member_parts or not all(len(parts) > 1 for parts in member_parts):
        return ""
    first_parts = {parts[0] for parts in member_parts}
    return next(iter(first_parts)) if len(first_parts) == 1 else ""


@dataclass
class PlanningStore:
    root: Path = DEFAULT_SCHEME_ROOT

    def __post_init__(self) -> None:
        self.root = Path(self.root).resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def scheme_dir(self, name: str) -> Path:
        clean = validate_scheme_name(name)
        path = (self.root / clean).resolve()
        if self.root not in path.parents and path != self.root:
            raise ValueError("方案路径越界")
        return path

    def workbook_path(self, name: str) -> Path:
        return self.scheme_dir(name) / WORKBOOK_NAME

    def time_series_workbook_path(self, name: str) -> Path:
        return self.scheme_dir(name) / TIME_SERIES_WORKBOOK_NAME

    def time_series_sqlite_path(self, name: str) -> Path:
        return self.scheme_dir(name) / TIME_SERIES_SQLITE_NAME

    def scheme_meta_path(self, name: str) -> Path:
        return self.scheme_dir(name) / SCHEME_META_NAME

    def read_scheme_meta(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        path = self.scheme_meta_path(clean)
        if not path.exists():
            return {}
        try:
            payload = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            return {}
        return payload if isinstance(payload, dict) else {}

    def write_scheme_meta(self, name: str, meta: dict[str, Any]) -> None:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        folder.mkdir(parents=True, exist_ok=True)
        path = folder / SCHEME_META_NAME
        path.write_text(json.dumps(meta, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")

    def scheme_owner_username(self, name: str) -> str:
        return str(self.read_scheme_meta(name).get("owner_username") or "").strip()

    def scheme_shared_usernames(self, name: str) -> list[str]:
        return normalize_shared_usernames(self.read_scheme_meta(name).get("shared_with_usernames"))

    def user_can_read_scheme(self, name: str, username: str) -> bool:
        clean_username = str(username or "").strip()
        if not clean_username:
            return False
        owner = self.scheme_owner_username(name)
        return owner == clean_username or clean_username in self.scheme_shared_usernames(name)

    def build_scheme_meta(self, owner_username: str) -> dict[str, Any]:
        owner = str(owner_username or "").strip()
        now = datetime.now().isoformat(timespec="seconds")
        return {
            "owner_username": owner,
            "created_by": owner,
            "created_at": now,
            "updated_at": now,
            "shared_with_usernames": [],
        }

    def list_schemes(
        self,
        owner_username: str | None = None,
        visible_username: str | None = None,
    ) -> list[dict[str, Any]]:
        owner_filter = None if owner_username is None else str(owner_username or "").strip()
        visible_filter = None if visible_username is None else str(visible_username or "").strip()
        schemes: list[dict[str, Any]] = []
        for folder in sorted(self.root.iterdir(), key=lambda item: item.name):
            if not folder.is_dir():
                continue
            workbook = folder / WORKBOOK_NAME
            time_series_workbook = folder / TIME_SERIES_WORKBOOK_NAME
            meta = self.read_scheme_meta(folder.name)
            owner = str(meta.get("owner_username") or "").strip()
            shared_with = normalize_shared_usernames(meta.get("shared_with_usernames"))
            if owner_filter is not None and owner != owner_filter:
                continue
            if visible_filter is not None and owner != visible_filter and visible_filter not in shared_with:
                continue
            modified_at = None
            if workbook.exists():
                modified_at = workbook.stat().st_mtime
            if time_series_workbook.exists():
                time_series_modified_at = time_series_workbook.stat().st_mtime
                modified_at = max(modified_at or time_series_modified_at, time_series_modified_at)
            schemes.append(
                {
                    "name": folder.name,
                    "has_workbook": workbook.exists(),
                    "modified_at": modified_at,
                    "owner_username": owner,
                    "created_by": str(meta.get("created_by") or "").strip(),
                    "created_at": str(meta.get("created_at") or "").strip(),
                    "shared_with_usernames": shared_with,
                }
            )
        return schemes

    def create_scheme(self, name: str, owner_username: str | None = None) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if folder.exists():
            raise FileExistsError(f"方案已存在: {clean}")
        folder.mkdir(parents=True)
        if owner_username is not None:
            self.write_scheme_meta(clean, self.build_scheme_meta(owner_username))
        self.write_scheme(clean, default_payload(clean))
        return self.read_scheme(clean)

    def copy_scheme(
        self,
        source: str,
        target: str,
        overwrite: bool = False,
        owner_username: str | None = None,
    ) -> dict[str, Any]:
        source_dir = self.scheme_dir(source)
        target_dir = self.scheme_dir(target)
        if not source_dir.exists():
            raise FileNotFoundError(f"源方案不存在: {source}")
        if target_dir.exists():
            if not overwrite:
                raise FileExistsError(f"目标方案已存在: {target}")
            file_ops.delete_directory_with_retry(target_dir, "目标方案目录")
        file_ops.copy_directory_with_retry(source_dir, target_dir, "方案目录")
        if owner_username is not None:
            self.write_scheme_meta(target, self.build_scheme_meta(owner_username))
        return self.read_scheme(target)

    def export_scheme_archive(self, name: str) -> bytes:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"方案不存在: {clean}")
        stream = BytesIO()
        with zipfile.ZipFile(stream, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(folder.rglob("*"), key=lambda item: item.relative_to(folder).as_posix()):
                if not path.is_file():
                    continue
                relative = path.relative_to(folder)
                if relative.name == SCHEME_META_NAME or relative.name.startswith("."):
                    continue
                archive.write(path, f"{clean}/{relative.as_posix()}")
        return stream.getvalue()

    def import_scheme_archive(
        self,
        content: bytes,
        target: str,
        owner_username: str | None = None,
        overwrite: bool = False,
    ) -> dict[str, Any]:
        clean = validate_scheme_name(target)
        target_dir = self.scheme_dir(clean)
        if target_dir.exists() and not overwrite:
            raise FileExistsError(f"方案已存在: {clean}")
        tmp_dir = Path(tempfile.mkdtemp(prefix=".scheme-import-", dir=self.root))
        try:
            self.extract_scheme_archive(content, tmp_dir)
            parameter_path = tmp_dir / WORKBOOK_NAME
            if not parameter_path.exists():
                raise ValueError(f"方案压缩包缺少{WORKBOOK_NAME}")
            self.validate_imported_workbook(parameter_path)
            time_series_path = tmp_dir / TIME_SERIES_WORKBOOK_NAME
            if time_series_path.exists():
                self.validate_imported_workbook(time_series_path)
            if target_dir.exists():
                file_ops.delete_directory_with_retry(target_dir, "已有方案目录")
            file_ops.replace_directory_with_retry(tmp_dir, target_dir, "导入方案目录")
            tmp_dir = target_dir
            if owner_username is not None:
                self.write_scheme_meta(clean, self.build_scheme_meta(owner_username))
            return self.read_scheme_overview(clean)
        finally:
            if tmp_dir.exists() and tmp_dir.name.startswith(".scheme-import-"):
                file_ops.delete_directory_with_retry(tmp_dir, "临时导入目录")

    def extract_scheme_archive(self, content: bytes, target_dir: Path) -> None:
        try:
            archive = zipfile.ZipFile(BytesIO(content))
        except zipfile.BadZipFile as exc:
            raise ValueError("方案压缩包无法读取") from exc
        with archive:
            member_infos = [info for info in archive.infolist() if not info.is_dir()]
            if len(member_infos) > SCHEME_ARCHIVE_MAX_FILES:
                raise ValueError("方案压缩包文件数量过多")
            total_size = sum(max(0, int(info.file_size or 0)) for info in member_infos)
            if total_size > SCHEME_ARCHIVE_MAX_UNCOMPRESSED_BYTES:
                raise ValueError("方案压缩包过大")
            member_parts = [safe_archive_member_parts(info.filename) for info in member_infos]
            strip_prefix = archive_strip_root_prefix([parts for parts in member_parts if parts])
            extracted = 0
            for info, parts in zip(member_infos, member_parts):
                if not parts:
                    continue
                relative_parts = parts[1:] if strip_prefix and parts[0] == strip_prefix else parts
                if not relative_parts:
                    continue
                if relative_parts[0] == "__MACOSX" or relative_parts[-1] in {".DS_Store", SCHEME_META_NAME}:
                    continue
                target_path = target_dir.joinpath(*relative_parts)
                resolved_target = target_path.resolve()
                if target_dir.resolve() not in resolved_target.parents and resolved_target != target_dir.resolve():
                    raise ValueError("方案压缩包包含非法路径")
                target_path.parent.mkdir(parents=True, exist_ok=True)
                with archive.open(info) as source, target_path.open("wb") as target_file:
                    shutil.copyfileobj(source, target_file)
                extracted += 1
            if extracted == 0:
                raise ValueError("方案压缩包为空")

    @staticmethod
    def validate_imported_workbook(path: Path) -> None:
        workbook = None
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
        except Exception as exc:
            raise ValueError(f"方案压缩包中的工作簿无法读取: {path.name}") from exc
        finally:
            if workbook is not None:
                workbook.close()

    def share_scheme(self, name: str, username: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"方案不存在: {clean}")
        target_username = str(username or "").strip()
        if not target_username:
            raise ValueError("分享用户名不能为空")
        meta = self.read_scheme_meta(clean)
        owner = str(meta.get("owner_username") or "").strip()
        if owner and target_username == owner:
            raise ValueError("不能将方案分享给创建者自己")
        shared_with = normalize_shared_usernames(meta.get("shared_with_usernames"))
        if target_username not in shared_with:
            shared_with.append(target_username)
        meta["shared_with_usernames"] = shared_with
        meta["updated_at"] = datetime.now().isoformat(timespec="seconds")
        self.write_scheme_meta(clean, meta)
        return self.read_scheme_meta(clean)

    def unshare_scheme(self, name: str, username: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"方案不存在: {clean}")
        target_username = str(username or "").strip()
        if not target_username:
            raise ValueError("取消分享用户名不能为空")
        meta = self.read_scheme_meta(clean)
        shared_with = [item for item in normalize_shared_usernames(meta.get("shared_with_usernames")) if item != target_username]
        meta["shared_with_usernames"] = shared_with
        meta["updated_at"] = datetime.now().isoformat(timespec="seconds")
        self.write_scheme_meta(clean, meta)
        return self.read_scheme_meta(clean)

    def rename_scheme(self, source: str, target: str) -> dict[str, Any]:
        source_dir = self.scheme_dir(source)
        target_dir = self.scheme_dir(target)
        if not source_dir.exists():
            raise FileNotFoundError(f"源方案不存在: {source}")
        if target_dir.exists():
            raise FileExistsError(f"目标方案已存在: {target}")
        file_ops.replace_directory_with_retry(source_dir, target_dir, "方案目录")
        return self.read_scheme(target)

    def delete_scheme(self, name: str) -> dict[str, str]:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        if not folder.exists() or not folder.is_dir():
            raise FileNotFoundError(f"方案不存在: {clean}")
        file_ops.delete_directory_with_retry(folder, "方案目录")
        return {"deleted": clean}

    def write_scheme(self, name: str, payload: dict[str, Any]) -> None:
        clean = validate_scheme_name(name)
        folder = self.scheme_dir(clean)
        folder.mkdir(parents=True, exist_ok=True)
        payload = sanitize_payload_names(deepcopy(payload))
        time_series_rows = payload.get("time_series") if "time_series" in payload else None
        workbook = build_workbook(payload | {"scheme": clean}, include_keys=[key for key in SHEET_SPECS if key != "time_series"])
        tmp_path = folder / f".{WORKBOOK_NAME}.tmp"
        final_path = folder / WORKBOOK_NAME
        try:
            file_ops.save_workbook_with_retry(workbook, tmp_path, "参数文件")
        finally:
            workbook.close()
        replace_workbook_with_retry(tmp_path, final_path)
        if "time_series" in payload and time_series_changed(self.time_series_workbook_path(clean), final_path, time_series_rows):
            time_series_workbook = build_time_series_workbook(payload | {"scheme": clean})
            time_series_tmp_path = folder / f".{TIME_SERIES_WORKBOOK_NAME}.tmp"
            time_series_final_path = folder / TIME_SERIES_WORKBOOK_NAME
            try:
                file_ops.save_workbook_with_retry(time_series_workbook, time_series_tmp_path, "8760时序数据文件")
            finally:
                time_series_workbook.close()
            replace_workbook_with_retry(time_series_tmp_path, time_series_final_path)
        if isinstance(time_series_rows, list):
            source_path = time_series_source_path(final_path)
            if should_refresh_time_series_sqlite(self.time_series_sqlite_path(clean), source_path):
                write_time_series_sqlite(self.time_series_sqlite_path(clean), source_path, time_series_rows)

    def read_scheme(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        path = self.workbook_path(clean)
        if not path.exists():
            raise FileNotFoundError(f"方案参数文件不存在: {path}")
        self.ensure_split_scheme_files(clean)
        payload = read_workbook(path, clean, include_keys=[key for key in SHEET_SPECS if key != "time_series"])
        payload.update(read_time_series_payload(path, clean))
        payload["validation"] = payload.get("validation", []) + validate_payload(payload)
        payload["time_series_loaded"] = True
        payload["time_series_count"] = len(payload.get("time_series", []))
        return payload

    def read_scheme_overview(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        path = self.workbook_path(clean)
        if not path.exists():
            raise FileNotFoundError(f"方案参数文件不存在: {path}")
        self.ensure_split_scheme_files(clean)
        payload = read_workbook(path, clean, include_keys=[key for key in SHEET_SPECS if key != "time_series"])
        payload["time_series_loaded"] = False
        payload["time_series_count"] = count_time_series_rows(path)
        payload["validation"] = payload.get("validation", []) + validate_payload(payload, require_time_series=False)
        return payload

    def read_time_series(self, name: str) -> dict[str, Any]:
        clean = validate_scheme_name(name)
        path = self.workbook_path(clean)
        if not path.exists():
            raise FileNotFoundError(f"方案参数文件不存在: {path}")
        self.ensure_split_scheme_files(clean)
        payload = read_time_series_payload(path, clean)
        payload["time_series_loaded"] = True
        payload["time_series_count"] = len(payload.get("time_series", []))
        payload["validation"] = payload.get("validation", []) + validate_payload(payload)
        return payload

    def ensure_split_scheme_files(self, name: str) -> None:
        clean = validate_scheme_name(name)
        parameter_path = self.workbook_path(clean)
        time_series_path = self.time_series_workbook_path(clean)
        if time_series_path.exists() or not workbook_has_sheet(parameter_path, SHEET_SPECS["time_series"][0]):
            return
        payload = read_workbook(parameter_path, clean)
        self.write_scheme(clean, payload)


def build_workbook(payload: dict[str, Any], include_keys: list[str] | None = None) -> Workbook:
    workbook = Workbook()
    workbook.remove(workbook.active)
    selected_keys = set(include_keys) if include_keys is not None else set(SHEET_SPECS)
    for key, (sheet_name, headers) in SHEET_SPECS.items():
        if key not in selected_keys:
            continue
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(headers)
        rows = payload.get(key, default_rows_for_key(key))
        if key == "planning_parameters" and isinstance(rows, dict):
            rows = [rows]
        for row in rows:
            sheet.append([row.get(header, field_default_for_key(key, header, "")) for header in headers])
    return workbook


def build_time_series_workbook(payload: dict[str, Any]) -> Workbook:
    return build_workbook(payload, include_keys=["time_series"])


def time_series_workbook_path_from_parameter_path(path: Path) -> Path:
    return path.with_name(TIME_SERIES_WORKBOOK_NAME)


def time_series_source_path(parameter_path: Path) -> Path:
    time_series_path = time_series_workbook_path_from_parameter_path(parameter_path)
    return time_series_path if time_series_path.exists() else parameter_path


def time_series_changed(time_series_path: Path, fallback_parameter_path: Path, rows: Any) -> bool:
    if not time_series_path.exists():
        return True
    try:
        existing_payload = read_time_series_workbook_payload(fallback_parameter_path, "")
    except Exception:
        return True
    existing_rows = existing_payload.get("time_series", [])
    return normalize_time_series_rows_for_compare(existing_rows) != normalize_time_series_rows_for_compare(rows)


def normalize_time_series_rows_for_compare(rows: Any) -> list[tuple[Any, ...]]:
    headers = SHEET_SPECS["time_series"][1]
    if not isinstance(rows, list):
        return []
    normalized = []
    for row in rows:
        if not isinstance(row, dict):
            normalized.append(tuple("" for _ in headers))
            continue
        normalized.append(tuple(row.get(header, field_default_for_key("time_series", header, "")) for header in headers))
    return normalized


class TimeSeriesSheetReadError(Exception):
    """Raised when only the 8760 time-series worksheet cannot be read."""


def read_workbook(path: Path, scheme: str, include_keys: list[str] | None = None) -> dict[str, Any]:
    selected_keys = set(include_keys) if include_keys is not None else set(SHEET_SPECS)
    variant = ("scheme_workbook", tuple(sorted(selected_keys)))
    return PARAMETER_WORKBOOK_CACHE.get(
        path,
        lambda resolved: read_workbook_uncached(resolved, scheme, selected_keys),
        variant=variant,
    )


def read_time_series_workbook_payload(parameter_path: Path, scheme: str) -> dict[str, Any]:
    source_path = time_series_source_path(parameter_path)
    return read_workbook(source_path, scheme, include_keys=["time_series"])


def read_time_series_payload(parameter_path: Path, scheme: str) -> dict[str, Any]:
    sqlite_payload = read_time_series_sqlite_payload(parameter_path, scheme)
    if sqlite_payload is not None:
        return sqlite_payload
    payload = read_time_series_workbook_payload(parameter_path, scheme)
    rows = payload.get("time_series")
    if isinstance(rows, list):
        try:
            write_time_series_sqlite(time_series_sqlite_path_from_parameter_path(parameter_path), time_series_source_path(parameter_path), rows)
        except (OSError, sqlite3.Error):
            pass
    return payload


def time_series_sqlite_path_from_parameter_path(path: Path) -> Path:
    return path.with_name(TIME_SERIES_SQLITE_NAME)


def read_time_series_sqlite_payload(parameter_path: Path, scheme: str) -> dict[str, Any] | None:
    db_path = time_series_sqlite_path_from_parameter_path(parameter_path)
    source_path = time_series_source_path(parameter_path)
    if not db_path.exists() or not source_path.exists():
        return None
    expected_signature = sqlite_source_signature(source_path)
    try:
        with sqlite3.connect(db_path) as connection:
            meta = read_sqlite_meta(connection)
            if meta.get("schema_version") != TIME_SERIES_SQLITE_SCHEMA_VERSION:
                return None
            if meta.get("source_signature") != expected_signature:
                return None
            rows = [
                {
                    "hour_index": hour_index,
                    "datetime": datetime_value,
                    "wind_speed": sqlite_display_value(wind_speed),
                    "solar_irradiance": sqlite_display_value(solar_irradiance),
                    "load": sqlite_display_value(load),
                    "temperature": sqlite_display_value(temperature),
                }
                for hour_index, datetime_value, wind_speed, solar_irradiance, load, temperature in connection.execute(
                    """
                    SELECT hour_index, datetime, wind_speed, solar_irradiance, load, temperature
                    FROM time_series
                    ORDER BY hour_index
                    """
                )
            ]
    except (OSError, sqlite3.Error):
        return None
    if not rows:
        return None
    return {"scheme": scheme, "validation": [], "time_series": rows}


def should_refresh_time_series_sqlite(db_path: Path, source_path: Path) -> bool:
    if not db_path.exists() or not source_path.exists():
        return True
    try:
        with sqlite3.connect(db_path) as connection:
            meta = read_sqlite_meta(connection)
        return (
            meta.get("schema_version") != TIME_SERIES_SQLITE_SCHEMA_VERSION
            or meta.get("source_signature") != sqlite_source_signature(source_path)
        )
    except (OSError, sqlite3.Error):
        return True


def write_time_series_sqlite(db_path: Path, source_path: Path, rows: list[dict[str, Any]]) -> None:
    if not isinstance(rows, list) or not source_path.exists():
        return
    db_path.parent.mkdir(parents=True, exist_ok=True)
    tmp_path = db_path.with_name(f".{db_path.name}.tmp")
    tmp_path.unlink(missing_ok=True)
    try:
        with sqlite3.connect(tmp_path) as connection:
            connection.execute("PRAGMA journal_mode=OFF")
            connection.execute("PRAGMA synchronous=OFF")
            connection.execute(
                """
                CREATE TABLE meta (
                    key TEXT PRIMARY KEY,
                    value TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                CREATE TABLE time_series (
                    hour_index INTEGER PRIMARY KEY,
                    datetime TEXT,
                    wind_speed REAL,
                    solar_irradiance REAL,
                    load REAL,
                    temperature REAL
                )
                """
            )
            connection.executemany(
                """
                INSERT INTO time_series (
                    hour_index,
                    datetime,
                    wind_speed,
                    solar_irradiance,
                    load,
                    temperature
                )
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                [
                    (
                        sqlite_int_value(row.get("hour_index")) or index,
                        str(row.get("datetime") or ""),
                        sqlite_float_value(row.get("wind_speed")),
                        sqlite_float_value(row.get("solar_irradiance")),
                        sqlite_float_value(row.get("load")),
                        sqlite_float_value(row.get("temperature")),
                    )
                    for index, row in enumerate(rows, start=1)
                    if isinstance(row, dict)
                ],
            )
            connection.executemany(
                "INSERT INTO meta (key, value) VALUES (?, ?)",
                [
                    ("schema_version", TIME_SERIES_SQLITE_SCHEMA_VERSION),
                    ("source_signature", sqlite_source_signature(source_path)),
                    ("row_count", str(len(rows))),
                    ("updated_at", datetime.now().isoformat(timespec="seconds")),
                ],
            )
        file_ops.replace_file_with_retry(tmp_path, db_path, "时序SQLite缓存")
    finally:
        tmp_path.unlink(missing_ok=True)


def read_sqlite_meta(connection: sqlite3.Connection) -> dict[str, str]:
    try:
        return {str(key): str(value) for key, value in connection.execute("SELECT key, value FROM meta")}
    except sqlite3.Error:
        return {}


def sqlite_source_signature(path: Path) -> str:
    stat = path.stat()
    return json.dumps({"mtime_ns": stat.st_mtime_ns, "size": stat.st_size}, sort_keys=True, separators=(",", ":"))


def sqlite_float_value(value: Any) -> float | int | None:
    if value in ("", None):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return int(number) if number.is_integer() else number


def sqlite_display_value(value: Any) -> Any:
    return "" if value is None else value


def sqlite_int_value(value: Any) -> int | None:
    number = sqlite_float_value(value)
    return int(number) if number is not None else None


def read_workbook_uncached(path: Path, scheme: str, selected_keys: set[str]) -> dict[str, Any]:
    repair_messages: list[dict[str, str]] = []
    if is_time_series_zip_member_corrupted(path):
        repair_messages.append(repair_time_series_sheet(path, "8760时序数据工作表压缩内容损坏"))
    for attempt in range(2):
        try:
            payload = read_workbook_once(path, scheme, selected_keys)
            payload["validation"] = repair_messages + payload.get("validation", [])
            return payload
        except TimeSeriesSheetReadError as exc:
            if attempt > 0:
                raise ValueError(f"参数文件损坏，8760时序数据工作表无法修复：{path}") from exc
            reason = str(exc)
            del exc
            gc.collect()
            repair_messages.append(repair_time_series_sheet(path, reason))
        except Exception as exc:
            if attempt == 0 and is_time_series_zip_member_corrupted(path):
                reason = str(exc)
                del exc
                gc.collect()
                repair_messages.append(repair_time_series_sheet(path, reason))
                continue
            raise ValueError(f"参数文件损坏或无法读取：{path}。请恢复备份或重新导入8760时序数据。原始错误：{exc}") from exc
    raise ValueError(f"参数文件损坏或无法读取：{path}")


def read_workbook_once(path: Path, scheme: str, selected_keys: set[str]) -> dict[str, Any]:
    workbook = load_workbook(path, data_only=True, read_only=True)
    payload: dict[str, Any] = {"scheme": scheme, "validation": [], "capacity_limits": []}
    header_presence: dict[str, set[str]] = {}
    try:
        for key, (sheet_name, headers) in SHEET_SPECS.items():
            if key not in selected_keys:
                continue
            if sheet_name not in workbook.sheetnames:
                payload[key] = default_rows_for_key(key)
                if key != "planning_parameters":
                    payload["validation"].append({"level": "error", "message": f"缺少工作表: {sheet_name}"})
                continue
            sheet = workbook[sheet_name]
            try:
                header_values = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1), [])]
            except Exception as exc:
                if key == "time_series":
                    raise TimeSeriesSheetReadError(str(exc)) from exc
                raise
            has_named_header = any(value is not None for value in header_values)
            header_index = {
                str(value): index
                for index, value in enumerate(header_values)
                if value is not None and str(value) in headers
            }
            header_presence[key] = set(header_index)
            rows = []
            try:
                for values in sheet.iter_rows(min_row=2, values_only=True):
                    if values is None or all(value is None for value in values):
                        continue
                    row = {}
                    for index, header in enumerate(headers):
                        if header in header_index:
                            source_index = header_index[header]
                        elif has_named_header:
                            row[header] = field_default_for_key(key, header, "")
                            continue
                        else:
                            source_index = index
                        row[header] = (
                            values[source_index]
                            if source_index < len(values) and values[source_index] is not None
                            else field_default_for_key(key, header, "")
                        )
                    if key == "planning_parameters" and "optimization_time_limit_minutes" not in header_index:
                        legacy_index = next(
                            (
                                index
                                for index, value in enumerate(header_values)
                                if value is not None and str(value) == "optimization_time_limit_seconds"
                            ),
                            None,
                        )
                        if legacy_index is not None and legacy_index < len(values) and values[legacy_index] not in (None, ""):
                            row["optimization_time_limit_minutes"] = numeric(values[legacy_index], 3600) / 60
                    if key == "planning_parameters" and not any(
                        field in header_index
                        for field in ("load_up_disturbance_factor", "load_down_disturbance_factor", "renewable_down_disturbance_factor")
                    ):
                        # Older schemes may still store a single load disturbance
                        # factor. Fan it out to the new split fields so those
                        # workbooks keep behaving sensibly after the schema change.
                        legacy_load_index = next(
                            (
                                index
                                for index, value in enumerate(header_values)
                                if value is not None and str(value) == "load_disturbance_factor"
                            ),
                            None,
                        )
                        if legacy_load_index is not None and legacy_load_index < len(values) and values[legacy_load_index] not in (None, ""):
                            legacy_load = numeric(values[legacy_load_index], 0.0)
                            row["load_up_disturbance_factor"] = legacy_load
                            row["load_down_disturbance_factor"] = legacy_load
                            row["renewable_down_disturbance_factor"] = 0.0
                    if key == "planning_parameters":
                        for legacy_field in ("storage_charge_efficiency", "storage_discharge_efficiency"):
                            legacy_index = next(
                                (
                                    index
                                    for index, value in enumerate(header_values)
                                    if value is not None and str(value) == legacy_field
                                ),
                                None,
                            )
                            if legacy_index is not None and legacy_index < len(values) and values[legacy_index] not in (None, ""):
                                row[legacy_field] = values[legacy_index]
                    rows.append(row)
            except Exception as exc:
                if key == "time_series":
                    raise TimeSeriesSheetReadError(str(exc)) from exc
                raise
            if key == "planning_parameters":
                payload[key] = [normalize_planning_parameter_row(row) for row in (rows or default_rows_for_key(key))]
            else:
                payload[key] = rows
        migrate_legacy_storage_efficiency(payload, header_presence.get("storage_pcs", set()))
    finally:
        workbook.close()
    return payload


def migrate_legacy_storage_efficiency(payload: dict[str, Any], storage_pcs_headers: set[str]) -> None:
    rows = payload.get("storage_pcs")
    if not isinstance(rows, list) or not rows:
        return
    planning_row = first_planning_parameter_row(payload)
    legacy_charge = planning_row.get("storage_charge_efficiency", "")
    legacy_discharge = planning_row.get("storage_discharge_efficiency", "")
    for row in rows:
        if not isinstance(row, dict):
            continue
        if "storage_charge_efficiency" not in storage_pcs_headers and legacy_charge not in ("", None):
            row["storage_charge_efficiency"] = legacy_charge
        if "storage_discharge_efficiency" not in storage_pcs_headers and legacy_discharge not in ("", None):
            row["storage_discharge_efficiency"] = legacy_discharge
    if isinstance(payload.get("planning_parameters"), list):
        for row in payload["planning_parameters"]:
            if isinstance(row, dict):
                row.pop("storage_charge_efficiency", None)
                row.pop("storage_discharge_efficiency", None)


def is_time_series_zip_member_corrupted(path: Path) -> bool:
    try:
        sheet_member = time_series_sheet_member(path)
        return corrupted_zip_member(path) == sheet_member
    except Exception:
        return False


def corrupted_zip_member(path: Path) -> str | None:
    with zipfile.ZipFile(path) as archive:
        for name in archive.namelist():
            try:
                archive.read(name)
            except Exception:
                return name
    return None


def repair_time_series_sheet(path: Path, reason: str) -> dict[str, str]:
    sheet_member = time_series_sheet_member(path)
    backup_path = backup_corrupted_workbook(path)
    replacement_xml = build_time_series_sheet_xml().encode("utf-8")
    tmp_path = path.with_name(f".{path.name}.repairing")
    try:
        write_repaired_time_series_workbook(path, tmp_path, sheet_member, replacement_xml)
        replace_workbook_with_retry(tmp_path, path)
    except Exception:
        file_ops.delete_file_if_exists_with_retry(tmp_path, "参数文件修复临时文件")
        raise
    return {
        "level": "warn",
        "message": (
            "检测到参数文件中的8760时序数据工作表损坏，"
            "已保留其它参数，已重建默认8760时序数据；"
            f"原文件已备份为 {backup_path.name}。原始错误：{reason}"
        ),
    }


def write_repaired_time_series_workbook(source_path: Path, tmp_path: Path, sheet_member: str, replacement_xml: bytes) -> None:
    def write_zip() -> None:
        with zipfile.ZipFile(source_path, "r") as source, zipfile.ZipFile(tmp_path, "w", compression=zipfile.ZIP_DEFLATED) as target:
            replaced = False
            for info in source.infolist():
                if info.filename == sheet_member:
                    target.writestr(info, replacement_xml)
                    replaced = True
                    continue
                target.writestr(info, source.read(info.filename))
            if not replaced:
                raise ValueError(f"未找到8760时序数据工作表: {sheet_member}")

    file_ops.retry_file_operation(
        write_zip,
        f"参数文件修复临时文件被占用，无法写入：{tmp_path.name}。请关闭正在打开的文件或预览窗口后重试。",
    )


def replace_workbook_with_retry(source: Path, target: Path, attempts: int = 20, delay_seconds: float = 0.1) -> None:
    file_ops.retry_file_operation(
        lambda: source.replace(target),
        f"参数文件被占用，无法保存：{target.name}。请关闭正在打开该文件的 Excel 或预览窗口后重试。",
        attempts=attempts,
        delay_seconds=delay_seconds,
    )
    file_cache.invalidate_path(source)
    file_cache.invalidate_path(target)


def backup_corrupted_workbook(path: Path) -> Path:
    timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
    backup_path = path.with_name(f"{path.stem}.corrupt-{timestamp}{path.suffix}.bak")
    counter = 1
    while backup_path.exists():
        backup_path = path.with_name(f"{path.stem}.corrupt-{timestamp}-{counter}{path.suffix}.bak")
        counter += 1
    file_ops.copy_file_with_retry(path, backup_path, "参数文件备份")
    return backup_path


def time_series_sheet_member(path: Path) -> str:
    sheet_name = SHEET_SPECS["time_series"][0]
    with zipfile.ZipFile(path) as archive:
        workbook_root = ElementTree.fromstring(archive.read(WORKBOOK_XML))
        relation_id = ""
        for sheet in workbook_root.findall(f".//{{{MAIN_XML_NS}}}sheet"):
            if sheet.attrib.get("name") == sheet_name:
                relation_id = sheet.attrib.get(f"{{{REL_XML_NS}}}id", "")
                break
        if not relation_id:
            raise ValueError(f"未找到工作表: {sheet_name}")
        rels_root = ElementTree.fromstring(archive.read(WORKBOOK_RELS_XML))
        for relationship in rels_root.findall(f".//{{{PACKAGE_REL_XML_NS}}}Relationship"):
            if relationship.attrib.get("Id") == relation_id:
                target = relationship.attrib.get("Target", "")
                if not target:
                    break
                normalized = target.lstrip("/")
                if not normalized.startswith("xl/"):
                    normalized = f"xl/{normalized}"
                return normalized.replace("\\", "/")
    raise ValueError(f"未找到工作表关系: {sheet_name}")


def build_time_series_sheet_xml() -> str:
    headers = SHEET_SPECS["time_series"][1]
    rows = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>',
        f'<worksheet xmlns="{MAIN_XML_NS}">',
        '<dimension ref="A1:F8761"/>',
        '<sheetData>',
        f'<row r="1">{"".join(inline_cell(column_name(index), 1, header) for index, header in enumerate(headers, start=1))}</row>',
    ]
    for row_index, row in enumerate(default_time_series(), start=2):
        rows.append(
            f'<row r="{row_index}">'
            f'<c r="A{row_index}" t="n"><v>{row["hour_index"]}</v></c>'
            f'{inline_cell("B", row_index, row["datetime"])}'
            f'<c r="C{row_index}" t="n"><v>{row["wind_speed"]}</v></c>'
            f'<c r="D{row_index}" t="n"><v>{row["solar_irradiance"]}</v></c>'
            f'<c r="E{row_index}" t="n"><v>{row["load"]}</v></c>'
            f'<c r="F{row_index}" t="n"><v>{row["temperature"]}</v></c>'
            "</row>"
        )
    rows.extend(["</sheetData>", "</worksheet>"])
    return "".join(rows)


def inline_cell(column: str, row: int, value: Any) -> str:
    text = (
        str(value)
        .replace("&", "&amp;")
        .replace("<", "&lt;")
        .replace(">", "&gt;")
        .replace('"', "&quot;")
    )
    return f'<c r="{column}{row}" t="inlineStr"><is><t>{text}</t></is></c>'


def column_name(index: int) -> str:
    name = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def count_sheet_rows(path: Path, sheet_name: str) -> int:
    return SHEET_ROW_COUNT_CACHE.get(
        path,
        lambda resolved: count_sheet_rows_uncached(resolved, sheet_name),
        variant=("sheet_rows", sheet_name),
    )


def count_time_series_rows(parameter_path: Path) -> int:
    return count_sheet_rows(time_series_source_path(parameter_path), SHEET_SPECS["time_series"][0])


def workbook_has_sheet(path: Path, sheet_name: str) -> bool:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        return sheet_name in workbook.sheetnames
    finally:
        workbook.close()


def count_sheet_rows_uncached(path: Path, sheet_name: str) -> int:
    workbook = load_workbook(path, data_only=True, read_only=True)
    try:
        if sheet_name not in workbook.sheetnames:
            return 0
        sheet = workbook[sheet_name]
        if sheet.max_row and sheet.max_row > 1:
            return sheet.max_row - 1
        return sum(1 for _ in sheet.iter_rows(min_row=2, values_only=True))
    finally:
        workbook.close()


def is_non_negative_integer_value(value: Any) -> bool:
    if isinstance(value, bool) or value in ("", None):
        return False
    if isinstance(value, int):
        return value >= 0
    if isinstance(value, float):
        return value >= 0 and value.is_integer()
    text = str(value).strip()
    return bool(re.fullmatch(r"\d+", text))


def numeric(value: Any, default: float = 0.0) -> float:
    if value in (None, ""):
        return default
    try:
        number = float(value)
    except (TypeError, ValueError):
        return default
    return number if number == number and number not in (float("inf"), float("-inf")) else default


def format_ratio(value: float) -> str:
    return f"{value:.4f}".rstrip("0").rstrip(".")


def active_device_soc_window(
    payload: dict[str, Any],
    key: str,
    capacity_field: str,
    default_lower: float,
    default_upper: float,
) -> tuple[float, float] | None:
    lower: float | None = None
    upper: float | None = None
    capacity_rule = DEVICE_FIELD_RULES[capacity_field]
    for row in payload.get(key, []):
        if not isinstance(row, dict):
            continue
        if not (
            validate_device_field_value(row.get(capacity_field, ""), capacity_rule)
            and validate_device_field_value(row.get("quantity_upper", ""), DEVICE_FIELD_RULES["quantity_upper"])
            and validate_device_field_value(row.get("soc_upper", ""), DEVICE_FIELD_RULES["soc_upper"])
            and validate_device_field_value(row.get("soc_lower", ""), DEVICE_FIELD_RULES["soc_lower"])
        ):
            continue
        if numeric(row.get(capacity_field), 0.0) <= 0 or numeric(row.get("quantity_upper"), 0.0) <= 0:
            continue
        row_lower = numeric(row.get("soc_lower"), default_lower)
        row_upper = numeric(row.get("soc_upper"), default_upper)
        if row_upper < row_lower:
            continue
        lower = row_lower if lower is None else max(lower, row_lower)
        upper = row_upper if upper is None else min(upper, row_upper)
    if lower is None or upper is None:
        return None
    return lower, upper


def active_storage_battery_soc_window(payload: dict[str, Any]) -> tuple[float, float] | None:
    return active_device_soc_window(payload, "storage_battery_packs", "battery_capacity", 0.1, 0.9)


def active_hydrogen_tank_soc_window(payload: dict[str, Any]) -> tuple[float, float] | None:
    return active_device_soc_window(payload, "hydrogen_tanks", "hydrogen_tank_capacity", 0.15, 0.85)


def validate_device_field_value(value: Any, rule: dict[str, Any]) -> bool:
    # Keep browser-side and server-side numeric validation in sync by using the
    # same small rule dictionary for both paths.
    if rule.get("integer"):
        if not is_non_negative_integer_value(value):
            return False
        number = float(value)
    else:
        if isinstance(value, bool) or value in ("", None):
            return False
        try:
            number = float(value)
        except (TypeError, ValueError):
            return False
        if not number == number or number in (float("inf"), float("-inf")):
            return False
    if rule.get("positive") and number <= 0:
        return False
    if rule.get("non_negative") and number < 0:
        return False
    if "min" in rule and number < float(rule["min"]):
        return False
    if "max" in rule and number > float(rule["max"]):
        return False
    return True


def validate_payload(payload: dict[str, Any], require_time_series: bool = True) -> list[dict[str, str]]:
    # Validation is layered so row field errors, cross-field consistency, and
    # global parameter issues can be reported independently.
    messages: list[dict[str, str]] = []
    if require_time_series:
        time_series = payload.get("time_series", [])
        if len(time_series) != 8760:
            messages.append({"level": "error", "message": f"时序数据行数应为8760，当前为{len(time_series)}"})
        else:
            messages.append({"level": "ok", "message": "时序数据行数正确"})
    elif "time_series_count" in payload:
        messages.append({"level": "ok", "message": f"时序数据延迟加载，当前行数为{payload['time_series_count']}"})

    for key in DEFAULT_DEVICE_ROWS:
        for index, row in enumerate(payload.get(key, []), start=1):
            row_label = f"{SHEET_SPECS[key][0]}第{index}行"
            for field in SHEET_SPECS[key][1]:
                rule = DEVICE_FIELD_RULES.get(field)
                if rule and not validate_device_field_value(row.get(field, ""), rule):
                    messages.append({"level": "error", "message": f"{row_label}{rule['message']}"})
            if (
                validate_device_field_value(row.get("quantity_lower", ""), DEVICE_FIELD_RULES["quantity_lower"])
                and validate_device_field_value(row.get("quantity_upper", ""), DEVICE_FIELD_RULES["quantity_upper"])
                and float(row.get("quantity_lower")) > float(row.get("quantity_upper"))
            ):
                messages.append({"level": "error", "message": f"{SHEET_SPECS[key][0]}第{index}行数量上限不能小于数量下限"})
            if key in {"storage_battery_packs", "hydrogen_tanks"}:
                soc_upper = row.get("soc_upper", "")
                soc_lower = row.get("soc_lower", "")
                if (
                    validate_device_field_value(soc_upper, DEVICE_FIELD_RULES["soc_upper"])
                    and validate_device_field_value(soc_lower, DEVICE_FIELD_RULES["soc_lower"])
                    and float(soc_upper) < float(soc_lower)
                ):
                    messages.append({"level": "error", "message": f"{SHEET_SPECS[key][0]}第{index}行SOC上限不能小于SOC下限"})
    messages.extend(validate_planning_parameters(payload))
    return messages


def first_planning_parameter_row(payload: dict[str, Any]) -> dict[str, Any]:
    rows = payload.get("planning_parameters")
    if isinstance(rows, dict):
        return rows
    if isinstance(rows, list) and rows:
        return rows[0]
    return deepcopy(DEFAULT_PLANNING_PARAMETERS)


def validate_planning_parameters(payload: dict[str, Any]) -> list[dict[str, str]]:
    # The planning-parameter sheet is a one-row table, so this helper owns all
    # scalar range checks for optimization and security settings.
    row = first_planning_parameter_row(payload)
    messages: list[dict[str, str]] = []

    def number_in_range(key: str, label: str, minimum: float | None = None, maximum: float | None = None) -> float | None:
        value = row.get(key, "")
        try:
            number = float(value)
        except (TypeError, ValueError):
            messages.append({"level": "error", "message": f"{label}必须为数值"})
            return None
        if minimum is not None and number < minimum:
            messages.append({"level": "error", "message": f"{label}不能小于{minimum:g}"})
        if maximum is not None and number > maximum:
            messages.append({"level": "error", "message": f"{label}不能大于{maximum:g}"})
        return number

    def integer_in_range(key: str, label: str, minimum: int, maximum: int) -> int | None:
        number = number_in_range(key, label, minimum, maximum)
        if number is None:
            return None
        if not float(number).is_integer():
            messages.append({"level": "error", "message": f"{label}必须为整数"})
            return None
        return int(number)

    number_in_range("diesel_price", "柴油价格(万元/吨)", 0)
    for key, label in (
        ("diesel_minimum_on_hours", "柴发开机持续工作小时数下限"),
        ("diesel_minimum_off_hours", "柴发关机持续工作小时数下限"),
    ):
        hours = number_in_range(key, label, 0, 24)
        if hours is not None and not float(hours).is_integer():
            messages.append({"level": "error", "message": f"{label}必须为整数"})
    if not is_valid_operation_mode_value(row.get("operation_mode")):
        messages.append({"level": "error", "message": "工作模式必须为全年运行或度夏运行"})
    start_month = integer_in_range("winter_start_month", "冬季开始月份", 1, 12)
    start_day = integer_in_range("winter_start_day", "冬季开始日期", 1, 31)
    end_month = integer_in_range("winter_end_month", "冬季结束月份", 1, 12)
    end_day = integer_in_range("winter_end_day", "冬季结束日期", 1, 31)
    for month, day, label in (
        (start_month, start_day, "冬季开始日期"),
        (end_month, end_day, "冬季结束日期"),
    ):
        if month is None or day is None:
            continue
        month_days = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
        if day > month_days:
            messages.append({"level": "error", "message": f"{label}不能超过{month}月{month_days}日"})
    number_in_range("green_power_ratio_lower", "绿色电量占比下限(0.0-1.0)", 0, 1)
    time_limit = number_in_range("optimization_time_limit_minutes", "规划求解时间上限(分钟)", 10, 1440)
    if time_limit is not None and not float(time_limit).is_integer():
        messages.append({"level": "error", "message": "规划求解时间上限(分钟)必须为正整数"})
    preferred_solver = normalize_preferred_solver(row.get("preferred_solver"))
    if preferred_solver not in {"auto", "gurobi", "cplex", "mosek", "scipy"}:
        messages.append({"level": "error", "message": "优先求解器必须为auto、gurobi、cplex、mosek或scipy"})
    initial_storage_soc = number_in_range("initial_storage_soc_ratio", "初始电储SOC(0.0-1.0)", 0, 1)
    storage_soc_window = active_storage_battery_soc_window(payload)
    if initial_storage_soc is not None and storage_soc_window is not None:
        lower, upper = storage_soc_window
        if lower > upper:
            messages.append({"level": "error", "message": "储能电池组SOC范围不存在公共区间，初始电储SOC无法同时满足所有储能电池组"})
        elif initial_storage_soc < lower or initial_storage_soc > upper:
            messages.append(
                {
                    "level": "error",
                    "message": (
                        "初始电储SOC(0.0-1.0)必须位于储能电池组SOC范围"
                        f"{format_ratio(lower)}-{format_ratio(upper)}内"
                    ),
                }
            )
    initial_hydrogen_soc = number_in_range("initial_hydrogen_storage_ratio", "初始氢储SOC(0.0-1.0)", 0, 1)
    hydrogen_soc_window = active_hydrogen_tank_soc_window(payload)
    if initial_hydrogen_soc is not None and hydrogen_soc_window is not None:
        lower, upper = hydrogen_soc_window
        if lower > upper:
            messages.append({"level": "error", "message": "储氢罐SOC范围不存在公共区间，初始氢储SOC无法同时满足所有储氢罐"})
        elif initial_hydrogen_soc < lower or initial_hydrogen_soc > upper:
            messages.append(
                {
                    "level": "error",
                    "message": (
                        "初始氢储SOC(0.0-1.0)必须位于储氢罐SOC范围"
                        f"{format_ratio(lower)}-{format_ratio(upper)}内"
                    ),
                }
            )
    storage_balance_mode = str(row.get("storage_balance_mode", "daily")).strip().lower()
    if storage_balance_mode not in {
        "daily",
        "weekly",
        "monthly",
        "annual",
        "none",
        "日内",
        "日内平衡",
        "每日",
        "周内",
        "周内平衡",
        "每周",
        "月度",
        "月度平衡",
        "年度",
        "年度平衡",
        "不闭环",
        "无",
    }:
        messages.append({"level": "error", "message": "电储能平衡模式必须为daily、weekly、monthly、annual或none"})
    for key, label in (
        ("storage_frequency_regulation_enabled", "储能是否参与调频"),
        ("frequency_security_constraint_enabled", "是否考虑频率安全约束"),
        ("post_disturbance_power_balance_enabled", "是否考虑扰动后平衡约束"),
        ("renewable_n_1_enabled", "是否考虑新能源N-1"),
        ("renewable_disturbance_enabled", "是否考虑新能源扰动"),
        ("load_disturbance_enabled", "是否考虑负荷扰动"),
    ):
        flag = number_in_range(key, label, 0, 1)
        if flag is not None and not float(flag).is_integer():
            messages.append({"level": "error", "message": f"{label}必须为0或1"})
    number_in_range("load_up_disturbance_factor", "负荷向上扰动系数(0.0-0.5)", 0, 0.5)
    number_in_range("load_down_disturbance_factor", "负荷向下扰动系数(0.0-0.5)", 0, 0.5)
    number_in_range("renewable_down_disturbance_factor", "新能源向下扰动系数(0.0-0.5)", 0, 0.5)
    nominal_frequency = number_in_range("nominal_frequency_hz", "额定频率(Hz)", 45, 65)
    nadir_lower = number_in_range("frequency_nadir_lower_hz", "频率最低点下限(Hz)", 45, 65)
    peak_upper = number_in_range("frequency_peak_upper_hz", "频率最高点上限(Hz)", 45, 65)
    number_in_range("frequency_lower_security_margin_hz", "频率下限安全裕度(Hz)", 0, 2)
    number_in_range("frequency_upper_security_margin_hz", "频率上限安全裕度(Hz)", 0, 2)
    number_in_range("load_frequency_coefficient_d", "负荷频率系数D", 0, 20)
    number_in_range("rocof_upper_hz_per_s", "RoCoF上限(Hz/s)", 0.0001, 20)
    steady_lower = number_in_range("steady_state_frequency_lower_hz", "稳态频率下限(Hz)", 0, 65)
    steady_upper = number_in_range("steady_state_frequency_upper_hz", "稳态频率上限(Hz)", 0, 65)
    number_in_range("frequency_governor_time_constant_s", "频率等效调速时间常数T(s)", 0, 20)
    number_in_range("frequency_nadir_evaluation_duration_s", "频率Nadir评估时长(s)", 1, 200)
    nadir_samples = number_in_range("nadir_linearization_samples_per_axis", "Nadir线性化每轴采样点数", 2, 7)
    if nadir_samples is not None and not float(nadir_samples).is_integer():
        messages.append({"level": "error", "message": "Nadir线性化每轴采样点数必须为正整数"})
    number_in_range("nadir_linearization_interval_ratio", "Nadir线性化区间比例", 0.05, 1)
    number_in_range("network_synchronization_coefficient_base", "网络同步系数基值", -100, 100)
    number_in_range("network_synchronization_coefficient_slope", "网络同步系数斜率", -100, 100)
    number_in_range("network_synchronization_reference_load_kw", "网络同步系数基准负荷(kW)", 0)
    if truthy_flag(row.get("frequency_security_constraint_enabled")):
        load_factor_sum = 0.0
        renewable_factor = 0.0
        if truthy_flag(row.get("load_disturbance_enabled")):
            load_factor_sum = max(0.0, numeric(row.get("load_up_disturbance_factor"), 0.0)) + max(
                0.0,
                numeric(row.get("load_down_disturbance_factor"), 0.0),
            )
        if truthy_flag(row.get("renewable_disturbance_enabled")):
            renewable_factor = max(0.0, numeric(row.get("renewable_down_disturbance_factor"), 0.0))
        renewable_n_1_enabled = truthy_flag(row.get("renewable_n_1_enabled"))
        if load_factor_sum <= 0 and renewable_factor <= 0 and not renewable_n_1_enabled:
            messages.append(
                {
                    "level": "warn",
                    "message": "频率安全约束已启用，但有效扰动系数均为0，频率波动曲线将保持在额定频率附近",
                }
            )
    if nominal_frequency is not None and nadir_lower is not None and nadir_lower > nominal_frequency:
        messages.append({"level": "error", "message": "频率最低点下限(Hz)不能大于额定频率(Hz)"})
    if nominal_frequency is not None and peak_upper is not None and peak_upper < nominal_frequency:
        messages.append({"level": "error", "message": "频率最高点上限(Hz)不能小于额定频率(Hz)"})
    if nominal_frequency is not None and steady_lower is not None and steady_lower > nominal_frequency:
        messages.append({"level": "error", "message": "稳态频率下限(Hz)不能大于额定频率(Hz)"})
    if nominal_frequency is not None and steady_upper is not None and steady_upper < nominal_frequency:
        messages.append({"level": "error", "message": "稳态频率上限(Hz)不能小于额定频率(Hz)"})
    if steady_upper is not None and steady_lower is not None and steady_upper < steady_lower:
        messages.append({"level": "error", "message": "稳态频率上限(Hz)不能小于稳态频率下限(Hz)"})
    return messages
