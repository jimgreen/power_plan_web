import shutil
import struct
import sys
import unittest
import zipfile
from pathlib import Path

from openpyxl import Workbook


WEB_ROOT = Path(__file__).resolve().parents[1]
sys.path.insert(0, str(WEB_ROOT))

import planning_store


class PlanningStoreTest(unittest.TestCase):
    def setUp(self):
        self.tmp_dir = WEB_ROOT / "tests" / "tmp_planning_store"
        shutil.rmtree(self.tmp_dir, ignore_errors=True)
        self.tmp_dir.mkdir(parents=True)
        self.store = planning_store.PlanningStore(root=self.tmp_dir)

    def tearDown(self):
        shutil.rmtree(self.tmp_dir, ignore_errors=True)

    def test_validate_scheme_name_accepts_chinese_letters_numbers(self):
        self.assertEqual(planning_store.validate_scheme_name("方案A-01"), "方案A-01")

    def test_validate_scheme_name_removes_whitespace_and_invisible_chars(self):
        self.assertEqual(planning_store.validate_scheme_name(" 方 案\r\nA\u200b "), "方案A")

    def test_validate_scheme_name_rejects_path_chars(self):
        for name in ("", "../bad", "a/b", "a\\b", ".", ".."):
            with self.assertRaises(ValueError):
                planning_store.validate_scheme_name(name)

    def test_create_scheme_writes_default_workbook(self):
        payload = self.store.create_scheme("方案A")

        workbook = self.tmp_dir / "方案A" / "parameters.xlsx"
        self.assertTrue(workbook.exists())
        self.assertEqual(payload["scheme"], "方案A")
        self.assertEqual(len(payload["time_series"]), 8760)
        self.assertIn("diesel_generators", payload)
        self.assertIn("storage_battery_packs", payload)
        self.assertIn("hydrogen_tanks", payload)
        self.assertIn("planning_parameters", payload)
        self.assertNotIn("design_life_years", payload["planning_parameters"][0])
        self.assertNotIn("planning_load_factor", payload["planning_parameters"][0])
        self.assertEqual(payload["planning_parameters"][0]["optimization_time_limit_minutes"], 60)
        self.assertEqual(payload["planning_parameters"][0]["initial_storage_soc_ratio"], 0.5)
        self.assertEqual(payload["planning_parameters"][0]["initial_hydrogen_storage_ratio"], 0.5)
        self.assertEqual(payload["planning_parameters"][0]["storage_charge_efficiency"], 0.95)
        self.assertEqual(payload["planning_parameters"][0]["storage_discharge_efficiency"], 0.95)
        self.assertEqual(payload["planning_parameters"][0]["storage_frequency_regulation_enabled"], 0)
        self.assertEqual(payload["planning_parameters"][0]["post_disturbance_power_balance_enabled"], 1)
        self.assertEqual(payload["planning_parameters"][0]["load_up_disturbance_factor"], 0)
        self.assertEqual(payload["planning_parameters"][0]["load_down_disturbance_factor"], 0)
        self.assertEqual(payload["planning_parameters"][0]["renewable_down_disturbance_factor"], 0)
        self.assertEqual(payload["validation"][0]["level"], "ok")
        for key in planning_store.DEFAULT_DEVICE_ROWS:
            self.assertIn("quantity_lower", payload[key][0])
            self.assertIn("quantity_upper", payload[key][0])
            self.assertNotIn("design_capacity_lower", payload[key][0])
            self.assertNotIn("design_capacity_upper", payload[key][0])
            self.assertEqual(payload[key][0]["quantity_lower"], 0)
            self.assertEqual(payload[key][0]["quantity_upper"], 0)
        self.assertNotIn("generation_efficiency", payload["photovoltaics"][0])
        self.assertNotIn("cut_in_wind_speed", payload["photovoltaics"][0])
        self.assertNotIn("cut_out_wind_speed", payload["photovoltaics"][0])
        self.assertIn("is_grid_forming", payload["storage_pcs"][0])
        self.assertEqual(payload["storage_pcs"][0]["is_grid_forming"], 0)
        self.assertEqual(payload["storage_battery_packs"][0]["soc_upper"], 0.9)
        self.assertEqual(payload["storage_battery_packs"][0]["soc_lower"], 0.1)

    def test_list_copy_and_rename_schemes(self):
        self.store.create_scheme("方案A")
        self.store.copy_scheme("方案A", "方案B")
        self.store.rename_scheme("方案B", "方案C")

        names = [item["name"] for item in self.store.list_schemes()]
        self.assertEqual(names, ["方案A", "方案C"])
        self.assertTrue((self.tmp_dir / "方案C" / "parameters.xlsx").exists())
        self.assertFalse((self.tmp_dir / "方案B").exists())

    def test_delete_scheme_removes_scheme_folder(self):
        self.store.create_scheme("方案A")
        self.store.create_scheme("方案B")

        result = self.store.delete_scheme("方案A")

        self.assertEqual(result["deleted"], "方案A")
        self.assertFalse((self.tmp_dir / "方案A").exists())
        self.assertTrue((self.tmp_dir / "方案B" / "parameters.xlsx").exists())

    def test_delete_scheme_rejects_missing_scheme(self):
        with self.assertRaises(FileNotFoundError):
            self.store.delete_scheme("不存在")

    def test_write_and_read_scheme_round_trip(self):
        self.store.create_scheme("方案A")
        payload = self.store.read_scheme("方案A")
        payload["time_series"][0]["wind_speed"] = 8.5
        payload["time_series"][0]["temperature"] = -12.5
        payload["diesel_generators"][0]["quantity_lower"] = 1
        payload["diesel_generators"][0]["quantity_upper"] = 3
        payload["hydrogen_tanks"][0]["hydrogen_tank_capacity"] = 300
        payload["storage_pcs"][0]["is_grid_forming"] = 1
        payload["storage_battery_packs"][0]["soc_upper"] = 0.85
        payload["storage_battery_packs"][0]["soc_lower"] = 0.15
        payload["planning_parameters"][0]["diesel_price"] = 0.76
        payload["planning_parameters"][0]["optimization_time_limit_minutes"] = 90
        payload["planning_parameters"][0]["initial_storage_soc_ratio"] = 0.25
        payload["planning_parameters"][0]["initial_hydrogen_storage_ratio"] = 0.75
        payload["planning_parameters"][0]["storage_charge_efficiency"] = 0.9
        payload["planning_parameters"][0]["storage_discharge_efficiency"] = 0.88
        payload["planning_parameters"][0]["storage_frequency_regulation_enabled"] = 1
        payload["planning_parameters"][0]["load_up_disturbance_factor"] = 0.1
        payload["planning_parameters"][0]["load_down_disturbance_factor"] = 0.2
        payload["planning_parameters"][0]["renewable_down_disturbance_factor"] = 0.3

        self.store.write_scheme("方案A", payload)
        saved = self.store.read_scheme("方案A")

        self.assertEqual(saved["time_series"][0]["wind_speed"], 8.5)
        self.assertEqual(saved["time_series"][0]["temperature"], -12.5)
        self.assertEqual(saved["diesel_generators"][0]["quantity_lower"], 1)
        self.assertEqual(saved["diesel_generators"][0]["quantity_upper"], 3)
        self.assertEqual(saved["hydrogen_tanks"][0]["hydrogen_tank_capacity"], 300)
        self.assertEqual(saved["storage_pcs"][0]["is_grid_forming"], 1)
        self.assertEqual(saved["storage_battery_packs"][0]["soc_upper"], 0.85)
        self.assertEqual(saved["storage_battery_packs"][0]["soc_lower"], 0.15)
        self.assertNotIn("generation_efficiency", saved["photovoltaics"][0])
        self.assertNotIn("design_life_years", saved["planning_parameters"][0])
        self.assertEqual(saved["planning_parameters"][0]["diesel_price"], 0.76)
        self.assertEqual(saved["planning_parameters"][0]["optimization_time_limit_minutes"], 90)
        self.assertEqual(saved["planning_parameters"][0]["initial_storage_soc_ratio"], 0.25)
        self.assertEqual(saved["planning_parameters"][0]["initial_hydrogen_storage_ratio"], 0.75)
        self.assertEqual(saved["planning_parameters"][0]["storage_charge_efficiency"], 0.9)
        self.assertEqual(saved["planning_parameters"][0]["storage_discharge_efficiency"], 0.88)
        self.assertEqual(saved["planning_parameters"][0]["storage_frequency_regulation_enabled"], 1)
        self.assertEqual(saved["planning_parameters"][0]["load_up_disturbance_factor"], 0.1)
        self.assertEqual(saved["planning_parameters"][0]["load_down_disturbance_factor"], 0.2)
        self.assertEqual(saved["planning_parameters"][0]["renewable_down_disturbance_factor"], 0.3)

    def test_read_scheme_overview_defers_time_series_rows(self):
        self.store.create_scheme("方案A")

        overview = self.store.read_scheme_overview("方案A")

        self.assertEqual(overview["scheme"], "方案A")
        self.assertNotIn("time_series", overview)
        self.assertFalse(overview["time_series_loaded"])
        self.assertEqual(overview["time_series_count"], 8760)
        self.assertIn("diesel_generators", overview)
        self.assertIn("planning_parameters", overview)
        self.assertEqual(overview["planning_parameters"][0]["frequency_security_upper"], 1.5)
        self.assertFalse(any(item["level"] == "error" for item in overview["validation"]))

    def test_read_time_series_returns_only_time_series_rows(self):
        self.store.create_scheme("方案A")

        payload = self.store.read_time_series("方案A")

        self.assertEqual(payload["scheme"], "方案A")
        self.assertEqual(len(payload["time_series"]), 8760)
        self.assertNotIn("diesel_generators", payload)
        self.assertNotIn("planning_parameters", payload)

    def test_read_scheme_repairs_corrupted_time_series_sheet_and_keeps_other_parameters(self):
        self.store.create_scheme("方案A")
        payload = self.store.read_scheme("方案A")
        payload["time_series"][0]["load"] = 88.8
        payload["diesel_generators"][0]["quantity_upper"] = 3
        payload["planning_parameters"][0]["diesel_price"] = 1.23
        self.store.write_scheme("方案A", payload)
        workbook_path = self.tmp_dir / "方案A" / "parameters.xlsx"
        corrupt_zip_member(workbook_path, "xl/worksheets/sheet1.xml")

        repaired = self.store.read_scheme("方案A")

        self.assertEqual(len(repaired["time_series"]), 8760)
        self.assertEqual(repaired["time_series"][0]["load"], 0)
        self.assertEqual(repaired["diesel_generators"][0]["quantity_upper"], 3)
        self.assertEqual(repaired["planning_parameters"][0]["diesel_price"], 1.23)
        self.assertTrue(
            any(
                item["level"] == "warn" and "8760时序数据" in item["message"] and "已重建" in item["message"]
                for item in repaired["validation"]
            )
        )
        self.assertTrue(any(workbook_path.parent.glob("parameters.corrupt-*.xlsx.bak")))
        overview = self.store.read_scheme_overview("方案A")
        self.assertEqual(overview["time_series_count"], 8760)
        reread = self.store.read_scheme("方案A")
        self.assertEqual(len(reread["time_series"]), 8760)
        self.assertFalse(any("已重建" in item["message"] for item in reread["validation"]))

    def test_default_time_series_includes_temperature(self):
        payload = planning_store.default_payload("方案A")

        self.assertIn("temperature", payload["time_series"][0])
        self.assertEqual(payload["time_series"][0]["temperature"], 0)

    def test_default_device_rows_include_design_life_years(self):
        payload = planning_store.default_payload("方案A")

        for key in planning_store.DEFAULT_DEVICE_ROWS:
            with self.subTest(key=key):
                self.assertIn("design_life_years", planning_store.SHEET_SPECS[key][1])
                self.assertEqual(payload[key][0]["design_life_years"], 20)

    def test_hydrogen_electrolyzer_rows_include_power_lower(self):
        payload = planning_store.default_payload("方案A")
        headers = planning_store.SHEET_SPECS["hydrogen_electrolyzers"][1]

        self.assertIn("power_lower", headers)
        self.assertLess(headers.index("power_capacity"), headers.index("power_lower"))
        self.assertLess(headers.index("power_lower"), headers.index("cost"))
        self.assertIn("power_lower", payload["hydrogen_electrolyzers"][0])
        self.assertEqual(payload["hydrogen_electrolyzers"][0]["power_lower"], 0)

    def test_storage_rows_include_grid_forming_and_soc_limits(self):
        payload = planning_store.default_payload("方案A")
        pcs_headers = planning_store.SHEET_SPECS["storage_pcs"][1]
        battery_headers = planning_store.SHEET_SPECS["storage_battery_packs"][1]

        self.assertIn("is_grid_forming", pcs_headers)
        self.assertLess(pcs_headers.index("quantity_upper"), pcs_headers.index("is_grid_forming"))
        self.assertLess(pcs_headers.index("is_grid_forming"), pcs_headers.index("design_life_years"))
        self.assertEqual(payload["storage_pcs"][0]["is_grid_forming"], 0)
        self.assertIn("soc_upper", battery_headers)
        self.assertIn("soc_lower", battery_headers)
        self.assertLess(battery_headers.index("battery_capacity"), battery_headers.index("soc_upper"))
        self.assertLess(battery_headers.index("soc_upper"), battery_headers.index("soc_lower"))
        self.assertLess(battery_headers.index("soc_lower"), battery_headers.index("cost"))
        self.assertEqual(payload["storage_battery_packs"][0]["soc_upper"], 0.9)
        self.assertEqual(payload["storage_battery_packs"][0]["soc_lower"], 0.1)

    def test_wind_turbine_rows_include_rated_wind_speed_between_cut_in_and_cut_out(self):
        payload = planning_store.default_payload("方案A")
        headers = planning_store.SHEET_SPECS["wind_turbines"][1]

        self.assertIn("rated_wind_speed", headers)
        self.assertLess(headers.index("cut_in_wind_speed"), headers.index("rated_wind_speed"))
        self.assertLess(headers.index("rated_wind_speed"), headers.index("cut_out_wind_speed"))
        self.assertIn("rated_wind_speed", payload["wind_turbines"][0])
        self.assertEqual(payload["wind_turbines"][0]["rated_wind_speed"], 12)

    def test_read_legacy_time_series_without_temperature_keeps_load(self):
        self.store.create_scheme("方案A")
        workbook_path = self.tmp_dir / "方案A" / "parameters.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        sheet = workbook.create_sheet("8760时序数据")
        sheet.append(["hour_index", "datetime", "wind_speed", "solar_irradiance", "load"])
        for hour in range(1, 8761):
            sheet.append([hour, f"H{hour:04d}", 8.5 if hour == 1 else 0, 310 if hour == 1 else 0, 123.4 if hour == 1 else 0])
        for key, (sheet_name, headers) in planning_store.SHEET_SPECS.items():
            if key == "time_series":
                continue
            device_sheet = workbook.create_sheet(sheet_name)
            device_sheet.append(headers)
        workbook.save(workbook_path)

        payload = self.store.read_scheme("方案A")

        self.assertEqual(payload["time_series"][0]["load"], 123.4)
        self.assertEqual(payload["time_series"][0]["temperature"], "")
        self.assertIn("planning_parameters", payload)
        self.assertNotIn("design_life_years", payload["planning_parameters"][0])

    def test_read_legacy_planning_parameters_defaults_missing_new_fields(self):
        self.store.create_scheme("方案A")
        workbook_path = self.tmp_dir / "方案A" / "parameters.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        time_sheet = workbook.create_sheet("8760时序数据")
        time_sheet.append(planning_store.SHEET_SPECS["time_series"][1])
        for row in planning_store.default_time_series():
            time_sheet.append([row.get(header, "") for header in planning_store.SHEET_SPECS["time_series"][1]])
        for key, (sheet_name, headers) in planning_store.SHEET_SPECS.items():
            if key in {"time_series", "planning_parameters"}:
                continue
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)
        planning_sheet = workbook.create_sheet("规划参数")
        planning_sheet.append(["diesel_price", "planning_load_factor", "green_power_ratio_lower"])
        planning_sheet.append([1.2, 1.1, 0.2])
        workbook.save(workbook_path)

        payload = self.store.read_scheme("方案A")
        row = payload["planning_parameters"][0]

        self.assertEqual(row["diesel_price"], 1.2)
        self.assertNotIn("planning_load_factor", row)
        self.assertEqual(row["green_power_ratio_lower"], 0.2)
        self.assertEqual(row["optimization_time_limit_minutes"], 60)
        self.assertEqual(row["initial_storage_soc_ratio"], 0.5)
        self.assertEqual(row["initial_hydrogen_storage_ratio"], 0.5)
        self.assertEqual(row["storage_charge_efficiency"], 0.95)
        self.assertEqual(row["storage_discharge_efficiency"], 0.95)
        self.assertEqual(row["load_up_disturbance_factor"], 0)
        self.assertEqual(row["load_down_disturbance_factor"], 0)
        self.assertEqual(row["renewable_down_disturbance_factor"], 0)

    def test_read_legacy_single_disturbance_factor_maps_to_split_fields(self):
        self.store.create_scheme("方案A")
        workbook_path = self.tmp_dir / "方案A" / "parameters.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        time_sheet = workbook.create_sheet("8760时序数据")
        time_sheet.append(planning_store.SHEET_SPECS["time_series"][1])
        for row in planning_store.default_time_series():
            time_sheet.append([row.get(header, "") for header in planning_store.SHEET_SPECS["time_series"][1]])
        for key, (sheet_name, headers) in planning_store.SHEET_SPECS.items():
            if key in {"time_series", "planning_parameters"}:
                continue
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(headers)
        planning_sheet = workbook.create_sheet("规划参数")
        planning_sheet.append(["diesel_price", "load_disturbance_factor"])
        planning_sheet.append([1.2, 0.12])
        workbook.save(workbook_path)

        payload = self.store.read_scheme("方案A")
        row = payload["planning_parameters"][0]

        self.assertEqual(row["load_up_disturbance_factor"], 0.12)
        self.assertEqual(row["load_down_disturbance_factor"], 0.12)
        self.assertEqual(row["renewable_down_disturbance_factor"], 0.0)

    def test_read_legacy_photovoltaic_sheet_does_not_shift_removed_capacity_columns(self):
        self.store.create_scheme("方案A")
        workbook_path = self.tmp_dir / "方案A" / "parameters.xlsx"
        workbook = Workbook()
        workbook.remove(workbook.active)
        time_sheet = workbook.create_sheet("8760时序数据")
        time_sheet.append(planning_store.SHEET_SPECS["time_series"][1])
        for row in planning_store.default_time_series():
            time_sheet.append([row.get(header, "") for header in planning_store.SHEET_SPECS["time_series"][1]])
        for key, (sheet_name, headers) in planning_store.SHEET_SPECS.items():
            if key in {"time_series", "photovoltaics"}:
                continue
            device_sheet = workbook.create_sheet(sheet_name)
            device_sheet.append(headers)
        pv_sheet = workbook.create_sheet("光伏参数")
        pv_sheet.append([
            "name",
            "capacity",
            "design_capacity_lower",
            "design_capacity_upper",
            "cost",
            "cut_in_wind_speed",
            "cut_out_wind_speed",
            "quantity_lower",
            "quantity_upper",
        ])
        pv_sheet.append(["旧光伏", 50, 10, 90, 3.5, 1, 2, 4, 5])
        workbook.save(workbook_path)

        payload = self.store.read_scheme("方案A")

        self.assertEqual(payload["photovoltaics"][0]["cost"], 3.5)
        self.assertNotIn("generation_efficiency", payload["photovoltaics"][0])
        self.assertEqual(payload["photovoltaics"][0]["design_life_years"], 20)
        self.assertEqual(payload["photovoltaics"][0]["quantity_lower"], 4)
        self.assertEqual(payload["photovoltaics"][0]["quantity_upper"], 5)

    def test_validate_quantity_limits(self):
        payload = planning_store.default_payload("方案A")
        payload["fuel_cells"][0]["quantity_lower"] = 5
        payload["fuel_cells"][0]["quantity_upper"] = 2

        messages = planning_store.validate_payload(payload)

        self.assertTrue(any("数据上限不能小于数据下限" in item["message"] for item in messages))

    def test_validate_device_numeric_field_rules(self):
        payload = planning_store.default_payload("方案A")
        payload["diesel_generators"][0]["quantity_lower"] = -1
        payload["diesel_generators"][0]["quantity_upper"] = 1.5
        payload["diesel_generators"][0]["design_life_years"] = 0
        payload["diesel_generators"][0]["cost"] = -0.1
        payload["diesel_generators"][0]["capacity"] = 0
        payload["diesel_generators"][0]["fuel_rate"] = 0
        payload["wind_turbines"][0]["cut_in_wind_speed"] = -0.1
        payload["wind_turbines"][0]["rated_wind_speed"] = 0
        payload["wind_turbines"][0]["cut_out_wind_speed"] = -0.2
        payload["storage_battery_packs"][0]["battery_capacity"] = 0
        payload["storage_battery_packs"][0]["soc_upper"] = 1.2
        payload["storage_battery_packs"][0]["soc_lower"] = -0.1
        payload["storage_pcs"][0]["is_grid_forming"] = 2
        payload["hydrogen_electrolyzers"][0]["electric_to_hydrogen_efficiency"] = 0
        payload["hydrogen_electrolyzers"][0]["power_lower"] = -0.1
        payload["fuel_cells"][0]["hydrogen_to_electric_efficiency"] = 0

        messages = planning_store.validate_payload(payload)
        message_text = "\n".join(item["message"] for item in messages)

        for expected in (
            "数据上下限必须为非负整数",
            "设计年限(年）必须为正整数",
            "成本(万元/台)必须为非负浮点数",
            "功率容量(kW)必须为正实数",
            "电池容量(kWh)必须为正实数",
            "SOC上限(0.0-1.0)必须在0到1之间",
            "SOC下限(0.0-1.0)必须在0到1之间",
            "是否构网必须为0或1",
            "电-氢效率(Nm3/kWh)必须为正实数",
            "氢-电效率(kWh/Nm3)必须为正实数",
            "油耗率(kg/kWh)必须为正实数",
            "功率下限(kW)必须为非负实数",
            "切入风速(m/s)必须为非负实数",
            "额定风速(m/s)必须为正实数",
            "切出风速(m/s)必须为非负实数",
        ):
            self.assertIn(expected, message_text)

    def test_write_scheme_sanitizes_invisible_device_names(self):
        payload = planning_store.default_payload("方案A")
        payload["diesel_generators"][0]["name"] = " 柴\u200b发\r\n1 "

        self.store.write_scheme("方案A", payload)
        saved = self.store.read_scheme("方案A")

        self.assertEqual(saved["diesel_generators"][0]["name"], "柴发1")

    def test_validate_time_series_messages_use_display_label(self):
        payload = planning_store.default_payload("方案A")
        payload["time_series"] = payload["time_series"][:12]

        messages = planning_store.validate_payload(payload)

        message_text = "\n".join(item["message"] for item in messages)
        self.assertIn("时序数据行数应为8760", message_text)
        self.assertNotIn("8760时序数据", message_text)

    def test_validate_planning_parameter_ranges(self):
        payload = planning_store.default_payload("方案A")
        payload["planning_parameters"][0]["green_power_ratio_lower"] = 1.2
        payload["planning_parameters"][0]["optimization_time_limit_minutes"] = 9
        payload["planning_parameters"][0]["initial_storage_soc_ratio"] = -0.1
        payload["planning_parameters"][0]["initial_hydrogen_storage_ratio"] = 1.1
        payload["planning_parameters"][0]["storage_charge_efficiency"] = 0
        payload["planning_parameters"][0]["storage_discharge_efficiency"] = 1.1
        payload["planning_parameters"][0]["load_up_disturbance_factor"] = -0.1
        payload["planning_parameters"][0]["load_down_disturbance_factor"] = 0.6
        payload["planning_parameters"][0]["renewable_down_disturbance_factor"] = "bad"
        payload["planning_parameters"][0]["frequency_security_upper"] = 1.1
        payload["planning_parameters"][0]["frequency_security_lower"] = 1.3

        messages = planning_store.validate_payload(payload)

        self.assertTrue(any("绿色电量占比下限(0.0-1.0)不能大于1" in item["message"] for item in messages))
        self.assertTrue(any("规划求解时间上限(分钟)不能小于10" in item["message"] for item in messages))
        self.assertTrue(any("初始电储SOC(0.0-1.0)不能小于0" in item["message"] for item in messages))
        self.assertTrue(any("初始氢储SOC(0.0-1.0)不能大于1" in item["message"] for item in messages))
        self.assertTrue(any("电储能充电效率(0.0-1.0)必须大于0" in item["message"] for item in messages))
        self.assertTrue(any("电储能放电效率(0.0-1.0)不能大于1" in item["message"] for item in messages))
        self.assertTrue(any("负荷向上扰动系数(0.0-0.5)不能小于0" in item["message"] for item in messages))
        self.assertTrue(any("负荷向下扰动系数(0.0-0.5)不能大于0.5" in item["message"] for item in messages))
        self.assertTrue(any("新能源向下扰动系数(0.0-0.5)必须为数值" in item["message"] for item in messages))
        self.assertTrue(any("频率安全上限不能小于频率安全下限" in item["message"] for item in messages))

    def test_validate_storage_soc_upper_not_below_lower(self):
        payload = planning_store.default_payload("方案A")
        payload["storage_battery_packs"][0]["soc_upper"] = 0.2
        payload["storage_battery_packs"][0]["soc_lower"] = 0.8

        messages = planning_store.validate_payload(payload)

        self.assertTrue(any("SOC上限不能小于SOC下限" in item["message"] for item in messages))


def corrupt_zip_member(path: Path, member_name: str) -> None:
    with zipfile.ZipFile(path) as archive:
        info = archive.getinfo(member_name)
        if info.compress_size <= 0:
            raise AssertionError(f"cannot corrupt empty member: {member_name}")
        data_offset = info.header_offset + 30
        with path.open("rb") as handle:
            handle.seek(info.header_offset + 26)
            name_length, extra_length = struct.unpack("<HH", handle.read(4))
            data_offset += name_length + extra_length
    with path.open("r+b") as handle:
        handle.seek(data_offset + min(10, info.compress_size - 1))
        original = handle.read(1)
        if not original:
            raise AssertionError(f"cannot read member byte: {member_name}")
        handle.seek(-1, 1)
        handle.write(bytes([original[0] ^ 0xFF]))


if __name__ == "__main__":
    unittest.main()
